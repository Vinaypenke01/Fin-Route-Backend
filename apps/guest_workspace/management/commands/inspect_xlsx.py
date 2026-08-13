import openpyxl
import json
from django.core.management.base import BaseCommand

class Command(BaseCommand):
    help = "Inspect Excel layout"

    def handle(self, *args, **options):
        excel_path = r"d:\Fintech\FinRoute_10_Borrowers_Correct_Date_Header_Payment_Row.xlsx"
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb.active

            rows = []
            for r in range(1, min(50, ws.max_row + 1)):
                row_vals = []
                for c in range(1, min(50, ws.max_column + 1)):
                    cell = ws.cell(row=r, column=c)
                    val = cell.value
                    fill_color = cell.fill.start_color.rgb if (cell.fill and cell.fill.start_color) else None
                    font_color = cell.font.color.rgb if (cell.font and cell.font.color) else None
                    is_bold = cell.font.bold if cell.font else False
                    
                    row_vals.append({
                        "col": c,
                        "val": str(val) if val is not None else "",
                        "fill": str(fill_color) if fill_color else "",
                        "font_color": str(font_color) if font_color else "",
                        "bold": is_bold
                    })
                non_empty = [x for x in row_vals if x["val"]]
                if non_empty:
                    max_col = max(x["col"] for x in non_empty)
                    rows.append({"row": r, "cells": row_vals[:max_col]})

            out_path = r"C:\Users\itzvi\.gemini\antigravity-ide\brain\467e738a-2c27-4476-8fbf-53213a1b5bef\scratch\excel_dump.json"
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(rows, f, indent=2)
            print(f"SUCCESS_DUMPED_{len(rows)}_ROWS")
        except Exception as e:
            print(f"ERROR_INSPECTING: {e}")
